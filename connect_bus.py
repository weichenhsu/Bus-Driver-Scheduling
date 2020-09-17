import openpyxl
from openpyxl import Workbook
import datetime

file_bus = '7'
sheet_names = ['0左', '0右', '1路', '2路', '3路', '5_1路', '5_2路', '6路', '7路', '9路', '10路', '11路', '14路', '15路', '18路', '19路', '20路', '21路', '77-1路', '77-2路', '88路', '99路', '99路賞鳥', '西環線']
station_id_0 = [0, 0, 0, 1, 3, 1, 1, 4, 0, 0, 0, 5, 0, 1, 0, 2, 0, 0, 2, 2, 0, 0, 6, 0]
station_id_1 = [0, 0, 7, 2, 0, 0, 4, 5, 3, 3, 3, 3, 2, 5, 3, 1, 3, 1, 2, 7, 0, 6, 6, 0]
excel_name_outbound = 'route_' + file_bus + '/公車資訊_謙慈_' + file_bus + '_0.xlsx'
excel_name_return = 'route_' + file_bus + '/公車資訊_謙慈_' + file_bus + '_1.xlsx'
data_outbound = []
data_return = []
bus_count_outbound = [] # 紀錄bus班次數量
bus_count_return = []
pljk = 0
# 去程
for num in range(0, len(sheet_names)):
    bus = []
    workbook = openpyxl.load_workbook(excel_name_outbound)
    worksheet = workbook[sheet_names[num]]       
    
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    bus_count_outbound.append(max_row - 3)

    for i in range(3, max_row): #跳掉前面三行直接抓發車資料
        row=[item.value for item in list(worksheet.rows)[i]]
        max_time = 0
        min_time = 24*60
        for j in range(1, max_col-1): # 搜尋該班次所有時間
            if isinstance(row[j], datetime.time):
                x = row[j].hour * 60 + row[j].minute
                if(x > max_time):
                    max_time = x
                if(x < min_time):
                    min_time = x
        if(max_time == 0 and min_time == 24*60): #沒有班次
            #print('沒有班次')
            #print(num, ' ', i, ' ', j)
            pljk = 1
        else:
            time_lag = max_time - min_time
            if(time_lag < 20):
                time_lag = 30
            if(num == 0 or num == 1 or num == 19 or num == 22): #0左, 0右, 88路, 西環線
                time_lag = 60
            max_time = min_time + time_lag
            new_row = [0, i-3, min_time, max_time, time_lag, station_id_0[num]] #[去程, 班次, 發車時間, 結束時間, 時間差]
            bus.append(new_row)
    data_outbound.append(bus)

#回程
for num in range(0, len(sheet_names)):
    bus = []
    workbook = openpyxl.load_workbook(excel_name_return)
    worksheet = workbook[sheet_names[num]]       
    
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    bus_count_outbound.append(max_row - 3)

    for i in range(3, max_row): #跳掉前面三行直接抓發車資料
        row=[item.value for item in list(worksheet.rows)[i]]
        max_time = 0
        min_time = 24*60
        for j in range(1, max_col-1): # 搜尋該班次所有時間
            if isinstance(row[j], datetime.time):
                x = row[j].hour * 60 + row[j].minute
                if(x > max_time):
                    max_time = x
                if(x < min_time):
                    min_time = x
        if(max_time == 0 and min_time == 24*60): #沒有班次
            #print('沒有班次')
            #print(num, ' ', i, ' ', j)
            #print(row)
            pljk = 1
        else:
            time_lag = max_time - min_time
            if(time_lag < 20):
                time_lag = 30
            if(num == 0 or num == 1 or num == 19 or num == 22): #0左, 0右, 88路, 西環線
                time_lag = 60
            max_time = min_time + time_lag
            new_row = [1, i-3, min_time, max_time, time_lag, station_id_1[num]] #[回程, 班次, 發車時間, 結束時間, 時間差]
            bus.append(new_row)            
    data_return.append(bus)

output = []
for num in range(0, len(sheet_names)):
    rows = []
    # 沒有車程
    if(len(data_outbound[num]) == 0 and len(data_return[num]) == 0):
        rows = []
        
    # 只有單趟
    elif(len(data_outbound[num]) == 0 or len(data_return[num]) == 0):
        # 對每班車進行輸出
        for i in range(0, len(data_outbound[num])):
            # [去程, 班次, 發車時間, 回程, 班次, 結束時間, 時間差, station]
            row = [0, i, data_outbound[num][i][2], -1, -1, data_outbound[num][i][3], data_outbound[num][i][4], data_outbound[num][i][5]]
            rows.append(row)
        for i in range(0, len(data_return[num])):
            row = [1, i, data_return[num][i][2], -1, -1, data_return[num][i][3], data_return[num][i][4], data_return[num][i][5]]
            rows.append(row)
        
    # 來回皆有
    else:
        # 儲存連接公車資料
        data_start = []
        data_continue = []
        start_id = 0
        continue_id = 1
        connect = []
        # 紀錄班次, 連接之後更改為0
        start_bus = []
        continue_bus = []
        if(data_outbound[num][0][2] <= data_return[num][0][2]): # 去程發車時間較回程早
            data_start = data_outbound[num]
            data_continue = data_return[num]
        else: # 去程發車時間較回程晚
            data_start = data_return[num]
            data_continue = data_outbound[num]
            start_id = 1
            continue_id = 0
        
        # 設定bus的班次，用以紀錄是否配對過
        for i in range(0, len(data_start)):
            start_bus.append(i)
        for i in range(0, len(data_continue)):
            continue_bus.append(i)
            
        # 連接班次    
        for i in range(0, len(data_start)):
            for j in range(0, len(data_continue)):
                if(data_start[i][3] < data_continue[j][2] and continue_bus[j] >= 0 and start_bus[i] >= 0):
                    if(data_continue[j][2] - data_start[i][3] <= 30):
                        row = [data_start[i][0], data_start[i][1], data_start[i][2], data_continue[j][0], data_continue[j][1], data_continue[j][3], data_continue[j][3] - data_start[i][2], data_start[i][5]]
                        start_bus[i] = -1
                        continue_bus[j] = -1;
                        rows.append(row)
        # 沒有連接到的班次
        for i in range(0, len(data_start)):
            if(start_bus[i] >= 0):
                row = [data_start[i][0], data_start[i][1], data_start[i][2], -1, -1, data_start[i][3] + data_start[i][4], data_start[i][4] * 2, data_start[i][5]]
                rows.append(row)
        for j in range(0, len(data_continue)):
            if(continue_bus[j] >= 0):
                row = [-1, -1, data_continue[j][2] - data_continue[j][4], data_continue[j][0], data_continue[j][1], data_continue[j][3], data_continue[j][4] * 2, data_start[0][5]]
                rows.append(row)
                
    output.append(rows)
    #print(output[num])
#print(output[5])
wrap = []
for num in range(0, len(sheet_names)):
    data = {}
    data['route_id'] = num
    data['route_name'] = sheet_names[num]
    data['content'] = []
    for i in range(len(output[num])):
        route_combine = []
        route_combine.append({
            'route_item_id':output[num][i][0], 
            'route_detail_id':output[num][i][1], 
        })
        route_combine.append({
            'route_item_id':output[num][i][3], 
            'route_detail_id':output[num][i][4], 
        })
        data['content'].append({
            'duration': output[num][i][6],
            'time_from': output[num][i][2], 
            'time_to': output[num][i][5], 
            'start_station':output[num][i][7], 
            'end_station':output[num][i][7], 
            'route_combine': route_combine
        })
    wrap.append(data)
import json
file_name = 'route-cal_'+file_bus+'.json'
with open(file_name, 'w', encoding='utf-8') as outfile:
    outfile.write(json.dumps(wrap, ensure_ascii=False, indent=4))